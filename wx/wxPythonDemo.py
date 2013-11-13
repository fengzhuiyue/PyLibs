#encoding=utf-8
#!/usr/bin/python

import os
import sys
import wx

ID_OPEN = 101
ID_EXIT = 110
ID_SAVE = 111
ID_BUTTON = 112


def excelParser(Dir,FileName):
    RequestFileName=Dir+'\\'+FileName
	print RequestFileName
	from openpyxl.reader.excel import load_workbook
	from openpyxl.workbook import Workbook	
	RequestFile = load_workbook(filename = r"'%s'" % RequestFileName)
	sheetnames = RequestFile.get_sheet_names()  
	ws = RequestFile.get_sheet_by_name(sheetnames[0])
	rowMax=ws.get_highest_row()-1

	SEC=[]
	MAQ=[]
	STL=[]
	IRI=[]

	for i in range(0,rowMax):
		SEC.append(str(ws.cell(row = i+1,column = 0).value))
		MAQ.append(int(ws.cell(row = i+1,column = 2).value))
		STL.append(int(ws.cell(row = i+1,column = 5).value))
		IRI.append(ws.cell(row = i+1,column = 10).value)

	import ConfigParser
	config=ConfigParser.ConfigParser()
	with open('Response.cfg','r') as cfgfile:
		config.readfp(cfgfile)
		#Choose a type
		ChooseType=config.get('Type','ChooseType')
		#Select fields
		Cfg_fromlegalEntityId=config.get(ChooseType,'fromlegalEntityId')
		Cfg_reclegalEntityId=config.get(ChooseType,'reclegalEntityId')
		Cfg_lendSecIdType=config.get(ChooseType,'lendSecIdType')
		Cfg_unitQty=config.get(ChooseType,'unitQty')
		Cfg_rateAmount=config.get(ChooseType,'rateAmount')
		Cfg_feeAmount=config.get(ChooseType,'feeAmount')
		Cfg_dividendRate=config.get(ChooseType,'dividendRate')
		Cfg_minimumFee=config.get(ChooseType,'minimumFee')
		Cfg_prePayRate=config.get(ChooseType,'prePayRate')
		Cfg_billingPrice=config.get(ChooseType,'billingPrice')
		Cfg_collType=config.get(ChooseType,'collType')
		Cfg_collDescription=config.get(ChooseType,'collDescription')
		Cfg_collCashAmount=config.get(ChooseType,'collCashAmount')
		Cfg_collCurrency=config.get(ChooseType,'collCurrency')
		Cfg_tradeDate=config.get(ChooseType,'tradeDate')
		Cfg_settlementDate=config.get(ChooseType,'settlementDate')
		Cfg_collateralDate=config.get(ChooseType,'collateralDate')

	import xlwt
	wbk = xlwt.Workbook()
	sheet = wbk.add_sheet('AutoBorrow Trading Records')
	List=['fromcorpId','fromsubAcctId','fromlegalEntityId','fromtraderId','reccorpId','recsubAcctId','reclegalEntityId','rectraderId','mesgType','mesgSubType','mesgNumber','mesgCount','borrowLoanIndicator','equilendId','equilendSeqNbr','lotGroupId','equilendAutoborrowBatchId','tradeType','tradeMethodType','lendSecId','lendSecIdType','borrSecId','borrSecIdType','unitQty','rateAmount','rateType','rateBenchmark','rateResetInterval','feeAmount','feeType','cashPaymentCurrency','cashPaymentCashAmount','dividendRate','taxVoucherDetails','MOD2','appendixC','cumOrExDividend','minimumFee','prePayRate','billingPrice','billingMargin','billingCurrency','billingCashAmount','billingFxRate','billDerivationInd','collType','collDescription','collMargin','collCashAmount','collContractPrice','collContractValue','collCurrency','collFxRate','tradeDate','settlementDate','collateralDate','termDate','termType','callableDate','settlementType','borrSettlementInstrId','lendSettlementInstrId','internalRefId','refAutoborrowBatchId','callableIndicator','callbackIndicator']
	ListLenth=len(List)
	for i in range(0,ListLenth) :
		sheet.write(0,i,List[i])

	lendSecIdType=[]
	lendSecIdType2=[]
	tradeDate=[]
	settlementDate=[]
	collateralDate=[]
	unitQty=[]
	for i in range(0,rowMax):
		if Cfg_tradeDate=='':
			tradeDate.append(STL[i])
		else:
			tradeDate.append(Cfg_tradeDate)
		if Cfg_settlementDate=='':
			settlementDate.append(STL[i])
		else:
			settlementDate.append(Cfg_settlementDate)
		if Cfg_unitQty=='':
			unitQty.append(str(MAQ[i]))
		else:
			unitQty.append(Cfg_unitQty)
		if Cfg_collateralDate=='':
			collateralDate.append(STL[i])
		else:
			collateralDate.append(Cfg_collateralDate)
			
		if len(SEC[i])==12:
			lendSecIdType2.append('I')
		elif len(SEC[i])==9:
			lendSecIdType2.append('C')
		elif len(SEC[i])==7:
			lendSecIdType2.append('S')
		else:
			lendSecIdType2.append('Error')
		
		if Cfg_lendSecIdType=='':
			lendSecIdType.append(lendSecIdType2[i])	
		else:
			lendSecIdType.append(Cfg_lendSecIdType)

		row=i+1
		sheet.write(row,2,Cfg_fromlegalEntityId)
		sheet.write(row,6,Cfg_reclegalEntityId)
		sheet.write(row,19,SEC[i])
		sheet.write(row,20,lendSecIdType[i])
		sheet.write(row,23,float(unitQty[i]))
		sheet.write(row,24,Cfg_rateAmount)
		sheet.write(row,28,Cfg_feeAmount)
		sheet.write(row,32,Cfg_dividendRate)
		sheet.write(row,37,Cfg_minimumFee)
		sheet.write(row,38,Cfg_prePayRate)
		sheet.write(row,39,Cfg_billingPrice)
		sheet.write(row,45,Cfg_collType)
		sheet.write(row,46,Cfg_collDescription)
		sheet.write(row,48,float(Cfg_collCashAmount))
		sheet.write(row,51,Cfg_collCurrency)
		sheet.write(row,53,int(tradeDate[i]))
		sheet.write(row,54,int(settlementDate[i]))
		sheet.write(row,55,int(collateralDate[i]))
		sheet.write(row,62,IRI[i])
		
	ResponseFileName='ResponseFile_%s.xls' %FileName

	wbk.save(ResponseFileName)



class MainWindow(wx.Frame):
    """ We simply derive a new class of Frame. """
    def __init__(self, parent, id, title):
        wx.Frame.__init__(self, parent, id, title, pos=(500,300), size=(900,500))
        self.SetMaxSize(wx.Size(600,450))
        self.SetMinSize(wx.Size(600,450))
        self.control = wx.TextCtrl(self, 1, style=wx.TE_MULTILINE)
        self.CreateStatusBar()
        filemenu  = wx.Menu()
        filemenu.Append(ID_OPEN,"&Open","open file")
        filemenu.AppendSeparator()
        filemenu.Append(ID_SAVE,"&Save","save file")
        filemenu.AppendSeparator()
        filemenu.Append(ID_EXIT,"&Exit","exit")
        menuBar = wx.MenuBar()
        menuBar.Append(filemenu,"&Menu")
        self.SetMenuBar(menuBar)
        wx.EVT_MENU(self,ID_OPEN,self.open)
        wx.EVT_MENU(self,ID_EXIT,self.exit)
        wx.EVT_MENU(self,ID_SAVE,self.save)
        self.sizer2 = wx.BoxSizer(wx.HORIZONTAL)
        self.buttons = []
       
#        for i in range(0,6):
#            self.buttons.append(wx.Button(self,ID_BUTTON+i,"Button &"+'i'))   
#            self.sizer2.Add(self.buttons[i],1,wx.EXPAND)
        self.sizer = wx.BoxSizer(wx.VERTICAL)
        self.sizer.Add(self.control,1,wx.EXPAND)
        self.sizer.Add(self.sizer2,0,wx.EXPAND)
        self.SetSizer(self.sizer)
        self.SetAutoLayout(1)
        self.sizer.Fit(self)
       
        self.Show(True)
        f = open('Response.cfg','r')
        self.control.SetValue(f.read())
        f.close()
       
    def exit(self,e):
        self.Close(True)
       
    def open(self,e):
        self.dirname = ''
        dlg = wx.FileDialog(self,"chose a file",self.dirname,"","*.*",wx.OPEN)
       
        if dlg.ShowModal() == wx.ID_OK:
            self.filename = dlg.GetFilename()
            self.dirname = dlg.GetDirectory()
            #f = open(os.path.join(self.dirname,self.filename),'r')
            #self.control.SetValue(f.read())
            #f.close()

            excelParser(self.dirname,self.filename)
        dlg.Destroy()
       
    def save(self,e):
        try:
            f = open('Response.cfg','w')
        except AttributeError:
            print 'File is not exist'
            sys.exit(0)
           
        content = self.control.GetValue()
        try:
            f.write(content)
        except UnboundLocalError:
            print 'File is not exist'
            sys.exit(0)
        finally:
            f.close()
                   
   
app = wx.PySimpleApp()
frame=MainWindow(None,-1, 'Equilend Simulator')
app.MainLoop()


